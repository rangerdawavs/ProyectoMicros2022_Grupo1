# lectura de color de casillas de excel

import openpyxl  # la libreria que usaremos

tabla_excel = openpyxl.load_workbook("Ejemplos_Rutas.xlsx", data_only=True)


def busca_color(color, celdas):
    element = 0
    result = list()
    for celda in celdas:
        element += 1
        if celda.fill.fgcolor.rgb == color:
            result.append(element)
    return result



##import openpyxl
##
##from openpyxl import load_workbook

##excel_file = 'Ejemplos_Rutas.xlsx'
##wb = load_workbook(excel_file, data_only = True)
##sh = wb['Modo 1']
##color_in_hex = sh['A1'].fill.start_color.index #valor en HEX
##print ('HEX =',color_in_hex) 
##print('RGB =', tuple(int(color_in_hex[i:i+2], 16) for i in (0, 2, 4))) # Color en RGB
